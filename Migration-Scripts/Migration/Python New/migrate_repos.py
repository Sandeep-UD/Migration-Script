#!/usr/bin/env python3
"""
GitHub Repository Migration Script using gh gei CLI

This script migrates repositories from a source organization to a target organization
using the GitHub Enterprise Importer (gh gei) CLI tool. It includes:
- Environment variable configuration
- Comprehensive logging
- Migration report generation in Excel format (only after successful migrations)
"""

import os
import sys
import csv
import time
import logging
import subprocess
import requests
import re
from datetime import datetime, timezone
from typing import Dict, List, Optional
from dataclasses import dataclass
from dotenv import load_dotenv
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------- Console & process encoding ----------
# Make stdout/stderr UTF-8 (prevents Windows cp1252 crashes on emoji/unicode)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
os.environ.setdefault("PYTHONIOENCODING", "utf-8")
os.environ.setdefault("PYTHONUTF8", "1")
# ------------------------------------------------

# Load environment variables
load_dotenv()

def safe_log_name(name: str) -> str:
    """Sanitize a string for use as a filename."""
    # Replace anything that's not alnum, dot, dash, or underscore
    return re.sub(r"[^A-Za-z0-9._-]+", "_", name)

@dataclass
class RepoInfo:
    """Data class to store repository information"""
    name: str
    org: str
    default_branch: str
    branches_count: int
    total_issues: int
    open_issues: int
    closed_issues: int
    total_prs: int
    open_prs: int
    closed_prs: int
    releases_count: int
    commits_count: int

class GitHubMigrator:
    """GitHub Repository Migrator using gh gei CLI"""
    
    def __init__(self):
        """Initialize the migrator with configuration from environment variables"""
        self.source_token = os.getenv('SOURCE_GITHUB_TOKEN')
        self.target_token = os.getenv('TARGET_GITHUB_TOKEN')
        self.source_org = os.getenv('SOURCE_ORGANIZATION')
        self.target_org = os.getenv('TARGET_ORGANIZATION')
        
        # Validate required environment variables
        self._validate_environment()
        
        # Setup logging
        self._setup_logging()
        
        # Migration report data
        self.migration_report = []
        self.successful_migrations = []
        self.repo_data = {
            'source_repos': [],
            'target_repos': []
        }
        
        self.logger.info("GitHub Migrator initialized successfully")
    
    def _validate_environment(self):
        """Validate that all required environment variables are set"""
        required_vars = [
            'SOURCE_GITHUB_TOKEN',
            'TARGET_GITHUB_TOKEN', 
            'SOURCE_ORGANIZATION',
            'TARGET_ORGANIZATION'
        ]
        
        missing_vars = [var for var in required_vars if not os.getenv(var)]
        
        if missing_vars:
            raise ValueError(f"Missing required environment variables: {', '.join(missing_vars)}")
    
    def _setup_logging(self):
        """Setup comprehensive logging configuration"""
        # Create logs directory if it doesn't exist
        self.logs_dir = Path('logs')
        self.logs_dir.mkdir(parents=True, exist_ok=True)
        
        # Configure logging with clean format
        log_filename = f"logs/migration_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        
        # Create custom formatter without module name
        class CleanFormatter(logging.Formatter):
            def format(self, record):
                log_time = datetime.fromtimestamp(record.created).strftime('%Y-%m-%d %H:%M:%S')
                # Ensure the message is properly encoded
                message = record.getMessage()
                # Replace any Unicode characters that might cause issues
                safe_message = message.encode('ascii', 'replace').decode('ascii')
                return f"[{log_time}] [{record.levelname}] {safe_message}"
        
        # Configure logging to only write to files (no console output)
        logging.basicConfig(
            level=logging.INFO,
            handlers=[
                logging.FileHandler(log_filename, encoding='utf-8')
            ]
        )
        
        # Add error-only logging similar to migrate_repos_old.py
        error_logger = logging.getLogger('ErrorLogger')
        error_handler = logging.FileHandler('migration_errors.log', encoding='utf-8')
        error_handler.setLevel(logging.ERROR)
        error_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        error_logger.addHandler(error_handler)
        error_logger.setLevel(logging.ERROR)
        
        # Apply custom formatter to file handler only
        for handler in logging.root.handlers:
            if isinstance(handler, logging.FileHandler):
                handler.setFormatter(CleanFormatter())
        
        self.logger = logging.getLogger('RepoMigrator')
        self.error_logger = error_logger
        self.log_filename = log_filename
    
    def _make_github_request(self, url: str, token: str, params: Dict = None) -> Optional[Dict]:
        """Make a GitHub API request with error handling"""
        headers = {
            'Authorization': f'token {token}',
            'Accept': 'application/vnd.github.v3+json'
        }
        
        max_retries = 3
        for attempt in range(max_retries):
            try:
                response = requests.get(url, headers=headers, params=params)
                
                if response.status_code == 200:
                    return response.json()
                elif response.status_code == 403:
                    self.logger.warning("Rate limit exceeded, waiting...")
                    time.sleep(60)  # Wait 1 minute
                    continue
                elif response.status_code == 404:
                    self.logger.warning(f"Resource not found: {url}")
                    return None
                else:
                    self.logger.error(f"API request failed: {response.status_code} - {response.text}")
                    return None
                    
            except Exception as e:
                self.logger.error(f"Request attempt {attempt + 1} failed: {e}")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)  # Exponential backoff
                else:
                    return None
        
        return None
    
    def _get_all_paginated_items(self, url: str, token: str, params: Dict = None) -> List:
        """Get all items from a paginated GitHub API endpoint"""
        all_items = []
        page = 1
        per_page = 100
        
        if params is None:
            params = {}
        
        while True:
            paginated_params = params.copy()
            paginated_params.update({'per_page': per_page, 'page': page})
            
            items = self._make_github_request(url, token, paginated_params)
            
            if not items or len(items) == 0:
                break
                
            all_items.extend(items)
            
            # If we got fewer items than per_page, we've reached the end
            if len(items) < per_page:
                break
                
            page += 1
            
            # Safety check to avoid infinite loops
            if page > 100:  # Max 10,000 items
                self.logger.warning(f"Reached pagination limit for {url}")
                break
        
        return all_items
    
    def _get_commits_count(self, org: str, repo_name: str, token: str, default_branch: str) -> int:
        """Get more accurate commit count for a repository"""
        try:
            # First try to get commits from the default branch
            commits_url = f"https://api.github.com/repos/{org}/{repo_name}/commits"
            
            # Make a HEAD request to get the total count from Link header
            headers = {
                'Authorization': f'token {token}',
                'Accept': 'application/vnd.github.v3+json'
            }
            
            # Get first page to check if there are commits
            params = {'sha': default_branch, 'per_page': 1}
            response = requests.get(commits_url, headers=headers, params=params)
            
            if response.status_code != 200:
                self.logger.warning(f"Could not fetch commits for {org}/{repo_name}, using fallback method")
                return self._get_commits_count_fallback(org, repo_name, token)
            
            # Check if there's a Link header for pagination
            link_header = response.headers.get('Link', '')
            if 'rel="last"' in link_header:
                # Extract the last page number from the Link header
                import re
                last_page_match = re.search(r'page=(\d+)[^>]*>;\s*rel="last"', link_header)
                if last_page_match:
                    last_page = int(last_page_match.group(1))
                    
                    # Get the last page to count remaining commits
                    last_page_params = {'sha': default_branch, 'per_page': 100, 'page': last_page}
                    last_response = requests.get(commits_url, headers=headers, params=last_page_params)
                    
                    if last_response.status_code == 200:
                        last_page_commits = len(last_response.json())
                        total_commits = (last_page - 1) * 100 + last_page_commits
                        return total_commits
            
            # If no pagination, count all commits on the first page
            first_page_commits = len(response.json())
            return first_page_commits
            
        except Exception as e:
            self.logger.warning(f"Error getting commit count for {org}/{repo_name}: {e}")
            return self._get_commits_count_fallback(org, repo_name, token)
    
    def _get_commits_count_fallback(self, org: str, repo_name: str, token: str) -> int:
        """Fallback method using contributors API for commit count"""
        try:
            commits_url = f"https://api.github.com/repos/{org}/{repo_name}/contributors"
            contributors_data = self._make_github_request(commits_url, token)
            return sum(contributor.get('contributions', 0) for contributor in contributors_data) if contributors_data else 0
        except:
            return 0
    
    def _get_repo_info(self, repo_name: str, org: str, token: str, is_target: bool = False) -> Optional[RepoInfo]:
        """Fetch comprehensive repository information"""
        self.logger.info(f"Fetching info for repository {org}/{repo_name}")
        
        # Get basic repo info
        repo_url = f"https://api.github.com/repos/{org}/{repo_name}"
        repo_data = self._make_github_request(repo_url, token)
        
        if not repo_data:
            self.logger.error(f"Failed to fetch repository data for {org}/{repo_name}")
            return None
        
        # Get branches
        branches_url = f"https://api.github.com/repos/{org}/{repo_name}/branches"
        branches_data = self._get_all_paginated_items(branches_url, token)
        branches_count = len(branches_data) if branches_data else 0
        
        # Get issues (account for migration issue in target repo)
        issues_url = f"https://api.github.com/repos/{org}/{repo_name}/issues"
        all_issues = self._get_all_paginated_items(issues_url, token, {'state': 'all'})
        open_issues = self._get_all_paginated_items(issues_url, token, {'state': 'open'})
        
        total_issues = len(all_issues) if all_issues else 0
        open_issues_count = len(open_issues) if open_issues else 0
        
        # Adjust for migration issue in target repository
        if is_target and total_issues > 0:
            # Check if there's a migration issue (usually titled "Migration Log", "Migration Issue" or similar)
            migration_issue_found = False
            migration_issue_is_open = False
            
            if all_issues:
                for issue in all_issues:
                    title = issue.get('title', '').lower()
                    if 'migration' in title:
                        migration_issue_found = True
                        # Check if this issue is in the open issues list
                        issue_number = issue.get('number')
                        if open_issues:
                            for open_issue in open_issues:
                                if open_issue.get('number') == issue_number:
                                    migration_issue_is_open = True
                                    break
                        break
            
            if migration_issue_found:
                total_issues -= 1
                if migration_issue_is_open:
                    open_issues_count -= 1
                self.logger.info(f"Adjusted issue count for {org}/{repo_name} (excluding migration issue)")
        
        closed_issues_count = total_issues - open_issues_count
        
        # Get pull requests
        pulls_url = f"https://api.github.com/repos/{org}/{repo_name}/pulls"
        all_prs = self._get_all_paginated_items(pulls_url, token, {'state': 'all'})
        open_prs = self._get_all_paginated_items(pulls_url, token, {'state': 'open'})
        
        total_prs = len(all_prs) if all_prs else 0
        open_prs_count = len(open_prs) if open_prs else 0
        closed_prs_count = total_prs - open_prs_count
        
        # Get releases
        releases_url = f"https://api.github.com/repos/{org}/{repo_name}/releases"
        releases_data = self._get_all_paginated_items(releases_url, token)
        releases_count = len(releases_data) if releases_data else 0
        
        # Get more accurate commits count using the commits API
        commits_count = self._get_commits_count(org, repo_name, token, repo_data.get('default_branch', 'main'))
        
        return RepoInfo(
            name=repo_name,
            org=org,
            default_branch=repo_data.get('default_branch', 'main'),
            branches_count=branches_count,
            total_issues=total_issues,
            open_issues=open_issues_count,
            closed_issues=closed_issues_count,
            total_prs=total_prs,
            open_prs=open_prs_count,
            closed_prs=closed_prs_count,
            releases_count=releases_count,
            commits_count=commits_count
        )
    
    def run_streaming(self, cmd, live_prefix="", log_path=None, env=None):
        """
        Run a shell command and stream stdout/stderr to console in real time.
        Also collects the full output and optionally writes to a log file.
        Returns (success: bool, combined_output: str)
        """
        # Prepare per-repo log file if requested
        log_file = None
        try:
            if log_path:
                os.makedirs(os.path.dirname(log_path), exist_ok=True)
                log_file = open(log_path, "w", encoding="utf-8")
        except Exception as e:
            print(f"[WARN] Could not open log file {log_path}: {e}", flush=True)

        # Merge stderr into stdout so we see everything
        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            bufsize=1,     # line-buffered
            shell=True,    # Fine from Git Bash / PowerShell; quoting is already added
            env=env
        )

        output_lines = []
        try:
            for line in proc.stdout:
                # Print each line as it arrives
                line = line.rstrip("\n")
                print(f"{live_prefix}{line}", flush=True)
                output_lines.append(line + "\n")
                if log_file:
                    log_file.write(line + "\n")

            proc.wait()
            success = (proc.returncode == 0)
        except KeyboardInterrupt:
            # If user hits Ctrl+C, stop the child process as well
            try:
                proc.terminate()
            except Exception:
                pass
            success = False
            output_lines.append("\n[ABORTED] Interrupted by user.\n")
            if log_file:
                log_file.write("\n[ABORTED] Interrupted by user.\n")
        finally:
            if log_file:
                log_file.close()

        return success, "".join(output_lines)

    def _run_gei_command(self, source_repo: str, target_repo: str = None) -> tuple:
        """Run the gh gei migrate-repo command using streaming output and per-repo logging"""
        try:
            # Set environment variables for the command
            env = os.environ.copy()
            env['GH_SOURCE_PAT'] = self.source_token
            env['GH_PAT'] = self.target_token
            
            # Use source repo name as target if not specified
            if not target_repo:
                target_repo = source_repo
            
            # Build command string for shell execution (similar to migrate_repos_old.py)
            command = (
                f'gh gei migrate-repo '
                f'--github-source-org "{self.source_org}" '
                f'--source-repo "{source_repo}" '
                f'--github-target-org "{self.target_org}" '
                f'--target-repo "{target_repo}" '
                f'--target-repo-visibility "private"'
            )
            
            self.logger.info(f"Starting migration: {self.source_org}/{source_repo} -> {self.target_org}/{target_repo}")
            
            # Create per-repo log file
            per_repo_log = self.logs_dir / f"{safe_log_name(source_repo)}__to__{safe_log_name(target_repo)}.log"
            
            # Use streaming execution with real-time output
            start_time = datetime.now(timezone.utc)
            success, output = self.run_streaming(
                command,
                live_prefix=f"[{source_repo} -> {target_repo}] ",
                log_path=str(per_repo_log),
                env=env
            )
            end_time = datetime.now(timezone.utc)
            
            duration_seconds = (end_time - start_time).total_seconds()
            
            if success:
                self.logger.info(f"Migration completed successfully: {self.source_org}/{source_repo} -> {self.target_org}/{target_repo} in {round(duration_seconds, 2)}s")
                return True, output, str(per_repo_log), start_time, end_time, duration_seconds
            else:
                self.logger.error(f"Migration failed for {self.source_org}/{source_repo} -> {self.target_org}/{target_repo}")
                self.error_logger.error(f"Migration failed for {source_repo} -> {target_repo}\n{output}")
                return False, output, str(per_repo_log), start_time, end_time, duration_seconds
                
        except Exception as e:
            self.logger.error(f"Error running migration command for {self.source_org}/{source_repo}: {e}")
            self.error_logger.error(f"Error running migration command for {source_repo}: {e}")
            return False, str(e), None, None, None, 0
    
    def migrate_repositories(self, csv_file_path: str):
        """Migrate repositories from CSV file"""
        self.logger.info(f"Starting repository migration from {csv_file_path}")
        
        try:
            with open(csv_file_path, 'r', newline='', encoding='utf-8-sig', errors='replace') as file:
                reader = csv.DictReader(file)
                repos_to_migrate = list(reader)
                total_repos = len(repos_to_migrate)
                
                self.logger.info(f"Found {total_repos} repositories to migrate")
                
                successful_migrations = 0
                failed_migrations = 0
                
                for index, row in enumerate(repos_to_migrate, 1):
                    source_repo = row['source_repo_name'].strip()
                    target_repo = row['target_repo_name'].strip()
                    
                    # Fallback: if target_repo is blank, reuse source_repo
                    if not target_repo:
                        target_repo = source_repo
                    
                    self.logger.info(f"Processing repository {index}/{total_repos}: {self.source_org}/{source_repo}")
                    print(f"[{index}/{total_repos}] Migrating: {self.source_org}/{source_repo} -> {self.target_org}/{target_repo}", flush=True)
                    
                    # Perform migration with detailed timing
                    migration_result = self._run_gei_command(source_repo, target_repo)
                    migration_success, output, log_file, start_time, end_time, duration_seconds = migration_result
                    
                    if migration_success:
                        successful_migrations += 1
                        self.successful_migrations.append({
                            'source_repo': source_repo,
                            'target_repo': target_repo
                        })
                        self.logger.info(f"Successfully migrated: {self.source_org}/{source_repo} -> {self.target_org}/{target_repo}")
                        print(f"✓ SUCCESS: {source_repo} migrated in {round(duration_seconds, 1)}s", flush=True)
                    else:
                        failed_migrations += 1
                        self.logger.error(f"Failed to migrate: {self.source_org}/{source_repo} -> {self.target_org}/{target_repo}")
                        print(f"✗ FAILED: {source_repo} migration failed", flush=True)
                    
                    # Add to migration report with detailed information
                    self.migration_report.append({
                        'Source Organization': self.source_org,
                        'Source Repository': source_repo,
                        'Target Organization': self.target_org,
                        'Target Repository': target_repo,
                        'Start Time': start_time.strftime('%Y-%m-%d %H:%M:%S') if start_time else 'N/A',
                        'End Time': end_time.strftime('%Y-%m-%d %H:%M:%S') if end_time else 'N/A',
                        'Duration (Seconds)': round(duration_seconds, 2) if duration_seconds else 0,
                        'Duration (Minutes)': round(duration_seconds / 60, 2) if duration_seconds else 0,
                        'Migration Status': 'Success' if migration_success else 'Failed'
                    })
                    
                    # No delay between migrations for faster processing
                
                self.logger.info(f"Migration process completed!")
                self.logger.info(f"All migrations finished. {successful_migrations + failed_migrations}/{total_repos} repos processed.")
                self.logger.info(f"Successful migrations: {successful_migrations}")
                self.logger.info(f"Failed migrations: {failed_migrations}")
                self.logger.info(f"Errors logged to: migration_errors.log")
                self.logger.info(f"Per-repo logs in: {self.logs_dir.resolve()}")
                
                # Console summary
                print(f"\nMigration Summary:", flush=True)
                print(f"  Total: {total_repos} repositories", flush=True)
                print(f"  Success: {successful_migrations}", flush=True)
                print(f"  Failed: {failed_migrations}", flush=True)
                
                return successful_migrations > 0
                
        except FileNotFoundError:
            self.logger.error(f"CSV file not found: {csv_file_path}")
            raise
        except Exception as e:
            self.logger.error(f"Error during migration: {e}")
            raise
    

    
    def collect_repo_data_for_successful_migrations(self):
        """Collect repository data only for successfully migrated repositories"""
        self.logger.info("Collecting repository data for successful migrations...")
        
        for migration in self.successful_migrations:
            source_repo = migration['source_repo']
            target_repo = migration['target_repo']
            
            self.logger.info(f"Collecting data for: {self.source_org}/{source_repo}")
            
            # Get source repository info
            source_info = self._get_repo_info(source_repo, self.source_org, self.source_token, is_target=False)
            if source_info:
                self.repo_data['source_repos'].append(source_info)
            
            # Wait a bit for the target repo to be fully accessible
            time.sleep(5)
            
            # Get target repository info
            target_info = self._get_repo_info(target_repo, self.target_org, self.target_token, is_target=True)
            if target_info:
                self.repo_data['target_repos'].append(target_info)
    
    def generate_migration_report(self, output_file: str = None):
        """Generate comprehensive migration report in Excel format (only if there were successful migrations)"""
        if not self.successful_migrations:
            self.logger.warning("No successful migrations found. Skipping report generation.")
            return
        
        if not output_file:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = f"migration_report_{timestamp}.xlsx"
        
        self.logger.info(f"Generating migration report: {output_file}")
        
        try:
            # Collect detailed repository data for successful migrations if not already collected
            if not self.repo_data['source_repos'] and not self.repo_data['target_repos']:
                self.collect_repo_data_for_successful_migrations()
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Migration Summary Sheet with timing information
                summary_df = pd.DataFrame(self.migration_report)
                summary_df.to_excel(writer, sheet_name='Migration_Summary', index=False)
                
                # Source Repositories Overview
                if self.repo_data['source_repos']:
                    source_data = []
                    for repo in self.repo_data['source_repos']:
                        source_data.append({
                            'Organization': repo.org,
                            'Repository': repo.name,
                            'Default Branch': repo.default_branch,
                            'Total Branches': repo.branches_count,
                            'Total Issues': repo.total_issues,
                            'Open Issues': repo.open_issues,
                            'Closed Issues': repo.closed_issues,
                            'Total PRs': repo.total_prs,
                            'Open PRs': repo.open_prs,
                            'Closed PRs': repo.closed_prs,
                            'Releases': repo.releases_count,
                            'Commits': repo.commits_count
                        })
                    
                    source_df = pd.DataFrame(source_data)
                    source_df.to_excel(writer, sheet_name='Source_Repositories', index=False)
                
                # Target Repositories Overview
                if self.repo_data['target_repos']:
                    target_data = []
                    for repo in self.repo_data['target_repos']:
                        target_data.append({
                            'Organization': repo.org,
                            'Repository': repo.name,
                            'Default Branch': repo.default_branch,
                            'Total Branches': repo.branches_count,
                            'Total Issues': repo.total_issues,
                            'Open Issues': repo.open_issues,
                            'Closed Issues': repo.closed_issues,
                            'Total PRs': repo.total_prs,
                            'Open PRs': repo.open_prs,
                            'Closed PRs': repo.closed_prs,
                            'Releases': repo.releases_count,
                            'Commits': repo.commits_count
                        })
                    
                    target_df = pd.DataFrame(target_data)
                    target_df.to_excel(writer, sheet_name='Target_Repositories', index=False)
                
                # Detailed comparison sheets
                self._create_detailed_sheets(writer)
            
            # Apply formatting
            self._format_excel_report(output_file)
            
            self.logger.info(f"Migration report generated successfully: {output_file}")
            
        except Exception as e:
            self.logger.error(f"Error generating migration report: {e}")
            raise
    
    def _create_detailed_sheets(self, writer):
        """Create detailed comparison sheets for different aspects"""
        # Create a mapping from source repo to target repo using successful migrations
        repo_mapping = {}
        for migration in self.successful_migrations:
            repo_mapping[migration['source_repo']] = migration['target_repo']
        
        # Branches comparison
        branches_data = []
        for source_repo in self.repo_data['source_repos']:
            # Find the corresponding target repo using the mapping
            target_repo_name = repo_mapping.get(source_repo.name)
            target_repo = next((r for r in self.repo_data['target_repos'] if r.name == target_repo_name), None)
            
            branches_data.append({
                'Repository': source_repo.name,
                'Source Org': source_repo.org,
                'Source Branches': source_repo.branches_count,
                'Source Default Branch': source_repo.default_branch,
                'Target Org': target_repo.org if target_repo else 'N/A',
                'Target Branches': target_repo.branches_count if target_repo else 'N/A',
                'Target Default Branch': target_repo.default_branch if target_repo else 'N/A',
                'Branches Match': 'Yes' if target_repo and source_repo.branches_count == target_repo.branches_count else 'No'
            })
        
        if branches_data:
            branches_df = pd.DataFrame(branches_data)
            branches_df.to_excel(writer, sheet_name='Branches_Comparison', index=False)
        
        # Issues comparison
        issues_data = []
        for source_repo in self.repo_data['source_repos']:
            target_repo_name = repo_mapping.get(source_repo.name)
            target_repo = next((r for r in self.repo_data['target_repos'] if r.name == target_repo_name), None)
            
            issues_data.append({
                'Repository': source_repo.name,
                'Source Total Issues': source_repo.total_issues,
                'Source Open Issues': source_repo.open_issues,
                'Source Closed Issues': source_repo.closed_issues,
                'Target Total Issues': target_repo.total_issues if target_repo else 'N/A',
                'Target Open Issues': target_repo.open_issues if target_repo else 'N/A',
                'Target Closed Issues': target_repo.closed_issues if target_repo else 'N/A',
                'Issues Match': 'Yes' if target_repo and source_repo.total_issues == target_repo.total_issues else 'No'
            })
        
        if issues_data:
            issues_df = pd.DataFrame(issues_data)
            issues_df.to_excel(writer, sheet_name='Issues_Comparison', index=False)
        
        # PRs comparison
        prs_data = []
        for source_repo in self.repo_data['source_repos']:
            target_repo_name = repo_mapping.get(source_repo.name)
            target_repo = next((r for r in self.repo_data['target_repos'] if r.name == target_repo_name), None)
            
            prs_data.append({
                'Repository': source_repo.name,
                'Source Total PRs': source_repo.total_prs,
                'Source Open PRs': source_repo.open_prs,
                'Source Closed PRs': source_repo.closed_prs,
                'Target Total PRs': target_repo.total_prs if target_repo else 'N/A',
                'Target Open PRs': target_repo.open_prs if target_repo else 'N/A',
                'Target Closed PRs': target_repo.closed_prs if target_repo else 'N/A',
                'PRs Match': 'Yes' if target_repo and source_repo.total_prs == target_repo.total_prs else 'No'
            })
        
        if prs_data:
            prs_df = pd.DataFrame(prs_data)
            prs_df.to_excel(writer, sheet_name='PRs_Comparison', index=False)
        
        # Releases comparison
        releases_data = []
        for source_repo in self.repo_data['source_repos']:
            target_repo_name = repo_mapping.get(source_repo.name)
            target_repo = next((r for r in self.repo_data['target_repos'] if r.name == target_repo_name), None)
            
            releases_data.append({
                'Repository': source_repo.name,
                'Source Releases': source_repo.releases_count,
                'Target Releases': target_repo.releases_count if target_repo else 'N/A',
                'Releases Match': 'Yes' if target_repo and source_repo.releases_count == target_repo.releases_count else 'No'
            })
        
        if releases_data:
            releases_df = pd.DataFrame(releases_data)
            releases_df.to_excel(writer, sheet_name='Releases_Comparison', index=False)
        
        # Commits comparison
        commits_data = []
        for source_repo in self.repo_data['source_repos']:
            target_repo_name = repo_mapping.get(source_repo.name)
            target_repo = next((r for r in self.repo_data['target_repos'] if r.name == target_repo_name), None)
            
            commits_data.append({
                'Repository': source_repo.name,
                'Source Commits': source_repo.commits_count,
                'Target Commits': target_repo.commits_count if target_repo else 'N/A',
                'Commits Match': 'Yes' if target_repo and source_repo.commits_count == target_repo.commits_count else 'No'
            })
        
        if commits_data:
            commits_df = pd.DataFrame(commits_data)
            commits_df.to_excel(writer, sheet_name='Commits_Comparison', index=False)
    
    def _format_excel_report(self, file_path: str):
        """Apply formatting to the Excel report"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = load_workbook(file_path)
            
            # Define styles
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply formatting to all sheets
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Format headers
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            
        except Exception as e:
            self.logger.warning(f"Could not apply Excel formatting: {e}")

def main():
    """Main function to run the migration script"""
    try:
        # Initialize migrator
        migrator = GitHubMigrator()
        
        # Check if gh gei CLI is installed
        try:
            result = subprocess.run(['gh', 'gei', '--version'], capture_output=True, text=True)
            if result.returncode != 0:
                raise subprocess.CalledProcessError(result.returncode, 'gh gei --version')
            version_info = result.stdout.strip()
            migrator.logger.info(f"GitHub Enterprise Importer CLI version: {version_info}")
        except (subprocess.CalledProcessError, FileNotFoundError):
            error_msg = "GitHub Enterprise Importer (gh gei) CLI not found. Please install it first: gh extension install github/gh-gei"
            migrator.logger.error(error_msg)
            print(f"ERROR: {error_msg}", flush=True)
            return
        
        # Check CSV file
        csv_file = 'repos.csv'
        if not os.path.exists(csv_file):
            error_msg = f"CSV file not found: {csv_file}"
            migrator.logger.error(error_msg)
            print(f"ERROR: {error_msg}", flush=True)
            return
        
        migrator.logger.info(f"Using CSV file: {csv_file}")
        
        # Perform migrations
        has_successful_migrations = migrator.migrate_repositories(csv_file)
        
        # Generate migration report only if there were successful migrations
        if has_successful_migrations:
            print("Generating migration report...", flush=True)
            migrator.logger.info("Generating migration report...")
            migrator.generate_migration_report()
            migrator.logger.info("Migration report generated successfully")
            print("Migration report generated successfully", flush=True)
        else:
            migrator.logger.warning("No successful migrations. Report generation skipped.")
            print("No successful migrations. Report generation skipped.", flush=True)
        
        print("Migration process completed!", flush=True)
        migrator.logger.info("Migration process completed successfully!")
        
    except KeyboardInterrupt:
        try:
            migrator.logger.info("Migration process interrupted by user")
        except:
            pass
        print("Migration interrupted by user", flush=True)
    except Exception as e:
        error_msg = f"Migration process failed: {e}"
        # Use the logger if available, otherwise fallback to basic logging
        try:
            migrator.logger.error(error_msg)
        except:
            logging.error(error_msg)
        print(f"ERROR: {error_msg}", flush=True)
        raise

if __name__ == "__main__":
    main()